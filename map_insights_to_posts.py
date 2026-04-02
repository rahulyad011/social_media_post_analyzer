"""
map_insights_to_posts.py
========================
Maps each insight to the top-N most relevant social media posts using:
  1. Bi-encoder (local)           — fast semantic similarity across all posts
  2. CrossEncoder reranking (opt) — more precise reranking of top candidates (config.yaml)
  3. LLM API (Anthropic or OpenAI) — scores and extracts exact excerpts from top candidates

Configuration : config.yaml
Secrets       : .env  (LLM_PROVIDER, ANTHROPIC_API_KEY, OPENAI_API_KEY)

Usage:
  python map_insights_to_posts.py
"""

# ──────────────────────────────────────────────────────────────
# Imports
# ──────────────────────────────────────────────────────────────

import os
import re
import sys
import json
import time
import logging
from types import SimpleNamespace

import numpy as np
import pandas as pd
import yaml
from dotenv import load_dotenv
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
# Logging  (configured before anything else so config-load errors appear)
# ──────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("run.log", mode="w", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# Suppress noisy third-party loggers (HuggingFace HTTP traffic, etc.)
for _noisy in ("httpx", "httpcore", "httpcore.http11", "sentence_transformers", "huggingface_hub"):
    logging.getLogger(_noisy).setLevel(logging.WARNING)

# ──────────────────────────────────────────────────────────────
# Config loading
# ──────────────────────────────────────────────────────────────

def _to_ns(obj):
    """Recursively convert dicts to SimpleNamespace for dot-access."""
    if isinstance(obj, dict):
        return SimpleNamespace(**{k: _to_ns(v) for k, v in obj.items()})
    if isinstance(obj, list):
        return [_to_ns(i) for i in obj]
    return obj


def load_config(path: str = "config.yaml") -> SimpleNamespace:
    if not os.path.exists(path):
        log.error(f"Config file not found: {path}")
        sys.exit(1)
    with open(path, encoding="utf-8") as fh:
        raw = yaml.safe_load(fh)
    cfg = _to_ns(raw)
    log.debug(f"Config loaded from {path}")
    return cfg


load_dotenv()
cfg = load_config()

# Convenience aliases so the rest of the code stays readable
COL_I  = cfg.columns.insights   # insight column names
COL_P  = cfg.columns.posts      # post column names
TUNE    = cfg.tuning
FILES   = cfg.files
MATCH   = cfg.matching


class _Models:
    """
    Model names — env vars take priority over config.yaml defaults.
      ANTHROPIC_MODEL  (default: models.anthropic in config.yaml)
      OPENAI_MODEL     (default: models.openai    in config.yaml)
    """
    anthropic: str = os.getenv("ANTHROPIC_MODEL", cfg.models.anthropic)
    openai:    str = os.getenv("OPENAI_MODEL",    cfg.models.openai)
    embedding: str = cfg.models.embedding


MODELS = _Models()

# Ordered list of insight columns written to output
ALL_INSIGHT_COLS = (
    [COL_I.num, COL_I.slide, COL_I.priority, COL_I.category, COL_I.text]
    + COL_I.stages
    + COL_I.stakeholders
)

# ──────────────────────────────────────────────────────────────
# Startup validation
# ──────────────────────────────────────────────────────────────

def validate():
    provider = os.getenv("LLM_PROVIDER", "anthropic").lower()

    log.info("=== Insight <-> Post Mapper (Vector + AI) ===")
    log.info(f"LLM provider : {provider}")

    if provider not in ("anthropic", "openai"):
        log.error(f"Invalid LLM_PROVIDER '{provider}'. Must be 'anthropic' or 'openai'.")
        sys.exit(1)

    model_name = MODELS.anthropic if provider == "anthropic" else MODELS.openai
    key_name   = "ANTHROPIC_API_KEY" if provider == "anthropic" else "OPENAI_API_KEY"
    key        = os.getenv(key_name, "")

    log.info(f"Model        : {model_name}")

    if not key:
        log.error(f"{key_name} is not set in .env")
        sys.exit(1)
    log.debug(f"{key_name} loaded (length={len(key)})")

    for path in (FILES.base, FILES.target):
        if not os.path.exists(path):
            log.error(f"Required file not found: {path}")
            sys.exit(1)
        log.debug(f"File found: {path}")

    log.info(f"Base file        : {FILES.base}   match column: '{MATCH.base_col}'")
    log.info(f"Target file      : {FILES.target}  match column: '{MATCH.target_col}'")

    ce = cfg.cross_encoder
    if ce.enabled:
        log.info(f"CrossEncoder     : ENABLED  model={ce.model}  top_k_retrieve={ce.top_k_retrieve}")
    else:
        log.info("CrossEncoder     : disabled")

    log.info(
        f"top_n={TUNE.top_n}  top_k_for_llm={TUNE.top_k_for_llm}  "
        f"min_similarity={TUNE.min_similarity}  min_score={TUNE.min_score}"
    )

    return provider

# ──────────────────────────────────────────────────────────────
# LLM client factory
# ──────────────────────────────────────────────────────────────

def build_client(provider: str):
    if provider == "openai":
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        log.debug("OpenAI client initialised")
    else:
        import anthropic
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        log.debug("Anthropic client initialised")
    return client

# ──────────────────────────────────────────────────────────────
# Step 1 — Load and clean data
# ──────────────────────────────────────────────────────────────

def load_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    log.info("Loading data...")

    insights_df = pd.read_excel(FILES.base,   dtype={COL_I.num: str})
    posts_df    = pd.read_excel(FILES.target, dtype={COL_P.id:  str})

    log.debug(f"Insights columns  : {list(insights_df.columns)}")
    log.debug(f"Posts columns     : {list(posts_df.columns)}")
    log.debug(f"Insights raw shape: {insights_df.shape}")
    log.debug(f"Posts raw shape   : {posts_df.shape}")

    # Clean insight columns
    for col in [COL_I.text, COL_I.priority, COL_I.category]:
        insights_df[col] = insights_df[col].fillna("").astype(str).str.strip()

    for col in COL_I.stages + COL_I.stakeholders:
        if col in insights_df.columns:
            insights_df[col] = insights_df[col].fillna("").astype(str).str.strip()
        else:
            log.debug(f"Optional insight column not present: '{col}'")

    # Clean post columns
    for col in (COL_P.text, COL_P.summary, COL_P.author_type):
        posts_df[col] = posts_df[col].fillna("").astype(str).str.strip()

    # Drop empty posts
    n_before = len(posts_df)
    posts_df = posts_df[posts_df[COL_P.text].str.len() > 0].reset_index(drop=True)
    log.debug(f"Dropped {n_before - len(posts_df)} empty posts")

    empty_insights = (insights_df[COL_I.text].str.len() == 0).sum()
    if empty_insights:
        log.warning(f"{empty_insights} insight(s) have empty text — they will produce no matches")

    log.info(f"Loaded {len(insights_df)} insights and {len(posts_df)} posts")
    return insights_df, posts_df

# ──────────────────────────────────────────────────────────────
# Step 2 — Embed posts and insights (runs locally, no API call)
# ──────────────────────────────────────────────────────────────

def embed_all(
    model: SentenceTransformer,
    insights_df: pd.DataFrame,
    posts_df: pd.DataFrame,
) -> tuple[np.ndarray, np.ndarray]:

    log.info(f"Embedding target posts — column: '{MATCH.target_col}'")
    post_embeddings = model.encode(
        posts_df[MATCH.target_col].tolist(), batch_size=64, show_progress_bar=True
    )
    log.debug(f"Post embeddings shape   : {post_embeddings.shape}")

    log.info(f"Embedding base insights — column: '{MATCH.base_col}'")
    insight_embeddings = model.encode(
        insights_df[MATCH.base_col].tolist(), batch_size=64, show_progress_bar=False
    )
    log.debug(f"Insight embeddings shape: {insight_embeddings.shape}")

    return insight_embeddings, post_embeddings

# ──────────────────────────────────────────────────────────────
# Step 3 — Vector similarity search
# ──────────────────────────────────────────────────────────────

def find_top_candidates(
    insight_embedding: np.ndarray,
    post_embeddings: np.ndarray,
    posts_df: pd.DataFrame,
) -> list[dict]:
    """
    Return bi-encoder candidates above min_similarity, sorted by similarity desc.
    Fetch limit: top_k_retrieve (CrossEncoder on) or top_k_for_llm (CrossEncoder off).
    """
    ce        = cfg.cross_encoder
    fetch_k   = ce.top_k_retrieve if ce.enabled else TUNE.top_k_for_llm
    sims      = cosine_similarity([insight_embedding], post_embeddings)[0]
    top_indices = np.argsort(sims)[::-1]

    candidates = []
    for idx in top_indices:
        sim = float(sims[idx])
        if sim < TUNE.min_similarity:
            break
        candidates.append({
            "post_id":      str(posts_df.iloc[idx][COL_P.id]),
            "post_text":    posts_df.iloc[idx][COL_P.text],
            "post_summary": posts_df.iloc[idx][COL_P.summary],
            "author_type":  posts_df.iloc[idx][COL_P.author_type],
            "similarity":   round(sim, 4),
        })
        if len(candidates) >= fetch_k:
            break

    log.debug(
        f"  Bi-encoder retrieval: {len(candidates)} candidates"
        f"  (fetch_k={fetch_k}, threshold={TUNE.min_similarity})"
    )
    for c in candidates:
        log.debug(
            f"    [{c['post_id']}] sim={c['similarity']:.4f}  author={c['author_type']}"
            f"  preview={c['post_text'][:80]!r}"
        )

    return candidates


# ──────────────────────────────────────────────────────────────
# Step 3b — CrossEncoder reranking (optional)
# ──────────────────────────────────────────────────────────────

def rerank_with_cross_encoder(
    cross_encoder,
    insight_text: str,
    candidates: list[dict],
) -> list[dict]:
    """
    Rerank candidates using the CrossEncoder, return top top_k_for_llm by CE score.
    Adds 'ce_score' to each candidate dict for debug visibility.
    """
    pairs  = [(insight_text, c["post_text"]) for c in candidates]
    scores = cross_encoder.predict(pairs)

    for c, score in zip(candidates, scores):
        c["ce_score"] = float(score)

    reranked = sorted(candidates, key=lambda x: x["ce_score"], reverse=True)

    log.debug(f"  CrossEncoder reranked {len(reranked)} → keeping top {TUNE.top_k_for_llm}")
    for c in reranked[: TUNE.top_k_for_llm]:
        log.debug(
            f"    [{c['post_id']}] ce_score={c['ce_score']:.4f}  sim={c['similarity']:.4f}"
            f"  author={c['author_type']}"
        )

    return reranked[: TUNE.top_k_for_llm]

# ──────────────────────────────────────────────────────────────
# Step 4 — LLM scoring and excerpt extraction
# ──────────────────────────────────────────────────────────────

_PROMPT_TEMPLATE = """\
You are an expert life sciences analyst reviewing social media posts about Alzheimer's disease (AD).

INSIGHT:
\"\"\"{insight_text}\"\"\"

For each post below:
1. Score its relevance to the insight on a scale of 0-10:
   - 10 = directly and explicitly supports or illustrates the insight
   - 7-9 = strongly related, same theme or patient/HCP/payer experience
   - 4-6 = partially related, tangentially mentions the topic
   - 1-3 = weak or incidental connection
   - 0 = completely unrelated

2. If score >= {min_score}, extract the specific excerpt (1-3 sentences maximum) from the post
   that most directly supports the insight. Do NOT paraphrase — copy exact words from the post.
   If score < {min_score}, set excerpt to "".

POSTS:
{posts_block}

Return ONLY a JSON array, no explanation, no markdown:
[
  {{"post_id": "<id>", "score": <0-10>, "excerpt": "<exact quote or empty string>"}},
  ...
]"""


def _call_llm(client, provider: str, prompt: str) -> str:
    """Call the appropriate LLM and return the raw text response."""
    if provider == "openai":
        response = client.chat.completions.create(
            model=MODELS.openai,
            max_tokens=3000,
            messages=[{"role": "user", "content": prompt}],
        )
        raw   = response.choices[0].message.content.strip()
        usage = response.usage
        log.debug(
            f"  OpenAI usage: prompt_tokens={usage.prompt_tokens}"
            f"  completion_tokens={usage.completion_tokens}"
        )
    else:
        response = client.messages.create(
            model=MODELS.anthropic,
            max_tokens=3000,
            messages=[{"role": "user", "content": prompt}],
        )
        raw   = response.content[0].text.strip()
        usage = response.usage
        log.debug(
            f"  Anthropic usage: input_tokens={usage.input_tokens}"
            f"  output_tokens={usage.output_tokens}"
        )

    log.debug(f"  LLM raw response ({len(raw)} chars):\n{raw[:500]}{'...' if len(raw) > 500 else ''}")
    return raw


def _parse_llm_response(raw: str) -> list[dict]:
    """Parse JSON from LLM response with a regex fallback."""
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        log.warning(f"  JSON parse failed ({e}), attempting regex fallback")
        match = re.search(r"\[.*\]", raw, re.DOTALL)
        if match:
            try:
                result = json.loads(match.group())
                log.debug("  Regex fallback succeeded")
                return result
            except json.JSONDecodeError as e2:
                log.error(f"  Regex fallback also failed: {e2}")
        else:
            log.error("  No JSON array found in LLM response")
        return []


def extract_excerpts(client, provider: str, insight_text: str, candidates: list[dict]) -> list[dict]:
    """
    Send top candidates to the LLM for scoring and excerpt extraction.
    Returns up to top_n results with score >= min_score, sorted by score desc.
    """
    if not candidates:
        log.debug("  No candidates — skipping LLM call")
        return []

    posts_block = "\n".join(f"[{c['post_id']}] {c['post_text']}" for c in candidates)
    prompt = _PROMPT_TEMPLATE.format(
        insight_text=insight_text,
        min_score=TUNE.min_score,
        posts_block=posts_block,
    )

    log.debug(f"  Sending {len(candidates)} candidates to {provider}")
    raw    = _call_llm(client, provider, prompt)
    scored = _parse_llm_response(raw)

    log.debug(f"  LLM returned {len(scored)} scored items")
    for item in scored:
        log.debug(
            f"    post_id={item.get('post_id')}  score={item.get('score')}"
            f"  excerpt_len={len(str(item.get('excerpt', '')))}"
        )

    # Merge LLM scores with candidate metadata; filter and rank
    id_to_candidate = {c["post_id"]: c for c in candidates}
    results = []
    for item in scored:
        pid     = str(item.get("post_id", ""))
        score   = float(item.get("score", 0))
        excerpt = str(item.get("excerpt", "")).strip()

        if pid not in id_to_candidate:
            log.warning(f"  LLM returned unknown post_id '{pid}' — skipping")
            continue
        if score < TUNE.min_score:
            log.debug(f"  post_id={pid} score={score} < min_score={TUNE.min_score} — excluded")
            continue

        c = id_to_candidate[pid]
        results.append({
            "post_id":      pid,
            "author_type":  c["author_type"],
            "post_summary": c["post_summary"],
            "excerpt":      excerpt,
            "similarity":   c["similarity"],
            "score":        score,
        })

    results.sort(key=lambda x: x["score"], reverse=True)
    final = results[: TUNE.top_n]

    log.debug(f"  Final matched posts ({len(final)}):")
    for p in final:
        log.debug(
            f"    [{p['post_id']}] score={p['score']}  sim={p['similarity']}"
            f"  author={p['author_type']}  excerpt={p['excerpt'][:80]!r}"
        )

    return final

# ──────────────────────────────────────────────────────────────
# Step 5 — Main mapping loop
# ──────────────────────────────────────────────────────────────

def map_insights(
    client,
    provider: str,
    insights_df: pd.DataFrame,
    posts_df: pd.DataFrame,
    insight_embeddings: np.ndarray,
    post_embeddings: np.ndarray,
    cross_encoder=None,
) -> list[dict]:

    n_total  = len(insights_df)
    no_match = 0
    results  = []

    for pos, (_, row) in enumerate(insights_df.iterrows()):
        insight_num  = str(row[COL_I.num])
        insight_text = row[COL_I.text]

        log.info(f"[{pos + 1}/{n_total}] Insight #{insight_num}: {insight_text[:80]}...")

        # Stage 1 — bi-encoder retrieval
        candidates = find_top_candidates(insight_embeddings[pos], post_embeddings, posts_df)
        log.info(f"  Bi-encoder candidates  : {len(candidates)}")

        # Stage 2 — CrossEncoder reranking (optional)
        if cross_encoder is not None and candidates:
            candidates = rerank_with_cross_encoder(cross_encoder, insight_text, candidates)
            log.info(f"  After CE reranking     : {len(candidates)}")

        # Stage 3 — LLM scoring + excerpt extraction
        top_posts = extract_excerpts(client, provider, insight_text, candidates)
        log.info(f"  Final matches          : {len(top_posts)} (score >= {TUNE.min_score})")

        if not top_posts:
            no_match += 1
            log.warning(f"  No posts matched for insight #{insight_num}")

        results.append({"row": row, "top_posts": top_posts})

        if candidates:
            time.sleep(TUNE.llm_sleep_sec)

    log.info(f"Mapping complete. {n_total - no_match}/{n_total} insights matched at least one post.")
    return results

# ──────────────────────────────────────────────────────────────
# Step 6 — Write results to Excel
# ──────────────────────────────────────────────────────────────

def _post_block_num(header: str) -> int:
    """Return the post rank (1-based) if header belongs to a post block, else 0."""
    for n in range(1, TUNE.top_n + 1):
        if header.startswith(f"Post {n} "):
            return n
    return 0


def _apply_formatting(ws, stage_headers: set, stakeholder_headers: set) -> None:
    xl = cfg.excel
    colors = xl.colors
    widths = xl.column_widths
    heights = xl.row_heights

    HEADER_FILL  = PatternFill("solid", fgColor=colors.header)
    INSIGHT_FILL = PatternFill("solid", fgColor=colors.insight)
    STAGE_FILL   = PatternFill("solid", fgColor=colors.stage)
    STAKEH_FILL  = PatternFill("solid", fgColor=colors.stakeholder)
    POST_FILLS   = [
        PatternFill("solid", fgColor=colors.post_even),
        PatternFill("solid", fgColor=colors.post_odd),
    ]
    THIN = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    # Header row
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN

    # Data rows
    center_headers = stage_headers | stakeholder_headers
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = str(ws.cell(row=1, column=cell.column).value or "")
            block  = _post_block_num(header)
            if block:
                cell.fill = POST_FILLS[(block - 1) % 2]
            elif header in stage_headers:
                cell.fill = STAGE_FILL
            elif header in stakeholder_headers:
                cell.fill = STAKEH_FILL
            else:
                cell.fill = INSIGHT_FILL
            cell.border    = THIN
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(
                wrap_text=True, vertical="top",
                horizontal="center" if header in center_headers else "left",
            )

    # Column widths
    for col in ws.columns:
        header = str(ws.cell(row=1, column=col[0].column).value or "")
        block  = _post_block_num(header)
        if header == COL_I.text:
            width = widths.insight_text
        elif header in stage_headers:
            width = widths.stage
        elif header in stakeholder_headers:
            width = widths.stakeholder
        elif header in (COL_I.num, COL_I.slide):
            width = widths.num_slide
        elif header == COL_I.priority:
            width = widths.priority
        elif header == COL_I.category:
            width = widths.category
        elif block and "Excerpt" in header:
            width = widths.post_excerpt
        elif block and "Summary" in header:
            width = widths.post_summary
        elif block and "Author" in header:
            width = widths.post_author
        elif block and "ID" in header:
            width = widths.post_id
        elif block and "Score" in header:
            width = widths.post_score
        else:
            width = widths.default
        ws.column_dimensions[get_column_letter(col[0].column)].width = width

    # Row heights
    ws.row_dimensions[1].height = heights.header
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = heights.data

    ws.freeze_panes = xl.freeze_pane


def write_output(insights_df: pd.DataFrame, results: list[dict]) -> None:
    log.info(f"Writing output to {FILES.output}...")

    output_rows = []
    for r in results:
        row  = r["row"]
        base = {col: row[col] for col in ALL_INSIGHT_COLS if col in insights_df.columns}

        for rank, post in enumerate(r["top_posts"], start=1):
            base[f"Post {rank} ID"]          = post["post_id"]
            base[f"Post {rank} Author Type"] = post["author_type"]
            base[f"Post {rank} Summary"]     = post["post_summary"]
            base[f"Post {rank} Excerpt"]     = post["excerpt"]
            base[f"Post {rank} Sim Score"]   = post["similarity"]
            base[f"Post {rank} AI Score"]    = round(post["score"], 1)

        output_rows.append(base)

    out_df = pd.DataFrame(output_rows)
    log.debug(f"Output dataframe shape: {out_df.shape}")
    out_df.to_excel(FILES.output, index=False)

    wb = load_workbook(FILES.output)
    _apply_formatting(wb.active, set(COL_I.stages), set(COL_I.stakeholders))
    wb.save(FILES.output)

    log.info(f"Output saved to: {FILES.output}")

# ──────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    provider = validate()

    insights_df, posts_df = load_data()

    log.info(f"Loading bi-encoder ({MODELS.embedding})...")
    embed_model = SentenceTransformer(MODELS.embedding)
    log.debug("Bi-encoder ready")

    insight_embeddings, post_embeddings = embed_all(embed_model, insights_df, posts_df)

    # Load CrossEncoder only if enabled
    cross_encoder = None
    if cfg.cross_encoder.enabled:
        from sentence_transformers import CrossEncoder
        log.info(f"Loading CrossEncoder ({cfg.cross_encoder.model})...")
        cross_encoder = CrossEncoder(cfg.cross_encoder.model)
        log.debug("CrossEncoder ready")

    client = build_client(provider)

    results = map_insights(
        client, provider, insights_df, posts_df,
        insight_embeddings, post_embeddings,
        cross_encoder=cross_encoder,
    )

    write_output(insights_df, results)
